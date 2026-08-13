# -*- coding: utf-8 -*-
"""同步总题库 xlsx 为 48 题修订版:
- 删除 6 道未试点题(hr-A7/A8、hr-B11/B12、zxy-B11/B12)
- 同步 43 道修订内容(模块A 18道 + 模块B 25道)
- 模块A 新增 keyFacts/mustFix/complianceHints 三列
输出: 测验题目_总题库(48题).xlsx (保留原 54 题文件不动)
"""
import json
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

bank = json.load(open('src/data/itemBank.json', encoding='utf-8'))
qa = {q['id']: q for q in bank['moduleA']['questions']}
qb = {q['id']: q for q in bank['moduleB']['questions']}

SRC = '测验题目_总题库(54题).xlsx'
DST = '测验题目_总题库(48题).xlsx'

wb = load_workbook(SRC)

# ---------- 模块A sheet ----------
wsA = wb['模块A_沉浸式交互任务(20题)']
wsA.title = '模块A_沉浸式交互任务(18题)'
# 清空数据行(保留表头)
wsA.delete_rows(2, wsA.max_row)
# 新表头(在原 10 列后追加 3 列)
headersA = ['题号', '业务场景类型', 'AI初稿状态', '风险等级', '任务背景与交付要求',
            'AI 生成初稿文本 (受测者可见)', '关键证据与事实基线 (需消耗点数查看)',
            '核心风险点/错漏分析', '关键核对与纠错动作', '标准修正策略与评分满分要点',
            '关键事实命中 keyFacts(评分)', '需删除错误 mustFix(评分)', '合规红线 complianceHints(评分)']
for c, h in enumerate(headersA, 1):
    wsA.cell(1, c, h)
    wsA.cell(1, c).font = Font(bold=True)
    wsA.cell(1, c).alignment = Alignment(wrap_text=True, vertical='center')

a_order = [q['id'] for q in bank['moduleA']['questions']]
for i, qid in enumerate(a_order, start=2):
    q = qa[qid]
    row = [q['id'], q['sceneType'], q['aiStatus'], q['riskLevel'], q['background'],
           q['aiDraft'], q['evidencePackage'], q['keyRisks'], q['keyCheckActions'],
           q['standardFixStrategy'],
           '、'.join(q.get('keyFacts') or []),
           '、'.join(q.get('mustFix') or []),
           '、'.join(q.get('complianceHints') or [])]
    for c, v in enumerate(row, 1):
        cell = wsA.cell(i, c, v)
        cell.alignment = Alignment(wrap_text=True, vertical='top')
    wsA.row_dimensions[i].height = 120

# 列宽
widthsA = [10, 12, 16, 10, 28, 34, 32, 30, 30, 34, 18, 18, 18]
for c, w in enumerate(widthsA, 1):
    wsA.column_dimensions[get_column_letter(c)].width = w

# ---------- 模块B sheet ----------
wsB = wb['模块B_微决策判断题(34题)']
wsB.title = '模块B_微决策判断题(30题)'
wsB.delete_rows(2, wsB.max_row)
headersB = ['题号', '场景分类', '对应核心维度', '微情境描述与人机冲突点',
            '最优动作 (2分)', '次优动作 (1分)', '无效动作 (0分)', '风险动作 (0分/-1分)', '心理测量学解析']
for c, h in enumerate(headersB, 1):
    wsB.cell(1, c, h)
    wsB.cell(1, c).font = Font(bold=True)
    wsB.cell(1, c).alignment = Alignment(wrap_text=True, vertical='center')

b_order = [q['id'] for q in bank['moduleB']['questions']]
for i, qid in enumerate(b_order, start=2):
    q = qb[qid]
    opt = {o['type']: o['text'] for o in q['options']}
    row = [q['id'], q['sceneType'], q['coreDimension'], q['scenario'],
           opt.get('best', ''), opt.get('good', ''), opt.get('invalid', ''), opt.get('risk', ''),
           q.get('analysis', '')]
    for c, v in enumerate(row, 1):
        cell = wsB.cell(i, c, v)
        cell.alignment = Alignment(wrap_text=True, vertical='top')
    wsB.row_dimensions[i].height = 80

widthsB = [10, 16, 16, 36, 36, 34, 32, 32, 34]
for c, w in enumerate(widthsB, 1):
    wsB.column_dimensions[get_column_letter(c)].width = w

# ---------- 输出 ----------
wb.save(DST)
print('已写出', DST)
print('模块A 行数(含表头):', wsA.max_row, '| 模块B 行数(含表头):', wsB.max_row)
print('模块A 题数:', len(a_order), '| 模块B 题数:', len(b_order))

# 校验
from openpyxl import load_workbook as lw2
chk = lw2(DST)
for name in chk.sheetnames:
    print(' sheet:', name)
mA = chk['模块A_沉浸式交互任务(18题)']
idsA = [mA.cell(r, 1).value for r in range(2, mA.max_row + 1)]
mB = chk['模块B_微决策判断题(30题)']
idsB = [mB.cell(r, 1).value for r in range(2, mB.max_row + 1)]
assert 'hr-A7' not in idsA and 'hr-A8' not in idsA, '模块A 删除失败!'
assert not any(x in idsB for x in ['hr-B11', 'hr-B12', 'zxy-B11', 'zxy-B12']), '模块B 删除失败!'
print('✓ 6 道未试点题均已移除')
print('✓ 表头新增:', [mA.cell(1, c).value for c in range(11, 14)])
print('✓ dsh-A1 keyFacts 列:', mA.cell(2, 11).value)
print('✓ zxy-A4 mustFix 列:', mA.cell(16, 12).value)
